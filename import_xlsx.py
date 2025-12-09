#!/usr/bin/env python3
"""Import data from Excel spreadsheets into QuoteForge database"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import load_workbook
from datetime import datetime, date
from app import app, db, Customer, Job

def parse_date(value):
    """Parse date from various formats"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y']
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except:
                continue
    return None

def parse_price(value):
    """Parse price from various formats"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Remove $ and commas
        value = value.replace('$', '').replace(',', '').strip()
        try:
            return float(value)
        except:
            return 0.0
    return 0.0

def import_sheet(ws, year, start_row=6, total_imported=0):
    """Import a single sheet"""
    imported = 0
    skipped = 0
    
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        # Column mapping: B=Date, C=Name, D=Address, E=Phone, F=pickup, G=Done, H=Description, I=QuoteNo, J=Price
        row_date = row[1] if len(row) > 1 else None
        name = row[2] if len(row) > 2 else None
        address = row[3] if len(row) > 3 else None
        phone = row[4] if len(row) > 4 else None
        description = row[7] if len(row) > 7 else None
        quote_num = row[8] if len(row) > 8 else None
        price = row[9] if len(row) > 9 else 0
        
        # Skip empty rows
        if not name and not description:
            continue
        
        # Clean up values
        name = str(name).strip() if name else f"Unknown {year}"
        phone = str(phone).strip() if phone else None
        if phone == '0' or phone == 'None':
            phone = None
        address = str(address).strip() if address else None
        if address == 'None':
            address = None
        description = str(description).strip() if description else "No description"
        
        job_date = parse_date(row_date)
        if not job_date:
            job_date = date(year, 1, 1)
        
        price_val = parse_price(price)
        
        # Find or create customer
        customer = None
        if phone:
            customer = Customer.query.filter(Customer.phone == phone).first()
        if not customer:
            customer = Customer.query.filter(Customer.name.ilike(name)).first()
        
        if not customer:
            customer = Customer(name=name, phone=phone, address=address)
            db.session.add(customer)
            db.session.flush()
        
        # Generate unique quote number
        base_num = total_imported + imported + 1
        qn = f"Q{base_num:05d}"
        
        # Ensure uniqueness
        while Job.query.filter(Job.quote_number == qn).first():
            base_num += 1
            qn = f"Q{base_num:05d}"
        
        # Check for duplicate
        existing = Job.query.filter(
            Job.customer_id == customer.id,
            Job.date == job_date,
            Job.description == description
        ).first()
        
        if existing:
            skipped += 1
            continue
        
        # Create job
        job = Job(
            customer_id=customer.id,
            quote_number=qn,
            description=description,
            price=price_val,
            date=job_date,
            status='completed'  # Old jobs are likely completed
        )
        db.session.add(job)
        imported += 1
        
        if imported % 50 == 0:
            db.session.commit()
            print(f"  Imported {imported} jobs...")
    
    db.session.commit()
    return imported, skipped

def main():
    xlsx_files = [
        "/home/bad/Desktop/David/quotes (version 1) (Autosaved) (Autosaved).xlsx",  # Nov 2025 data!
        "/home/bad/Desktop/David/quotes (version 1) (Autosaved) (Autosaved) (Autosaved).xlsx",
        "/home/bad/Desktop/David/quotes (version 1) (Autosaved).xlsx",
    ]
    
    with app.app_context():
        total_imported = 0
        total_skipped = 0
        
        # Use the first available file
        for xlsx_path in xlsx_files:
            if os.path.exists(xlsx_path):
                print(f"\nImporting from: {xlsx_path}")
                wb = load_workbook(xlsx_path, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    if sheet_name.isdigit():
                        year = int(sheet_name)
                        print(f"\n=== Importing {sheet_name} sheet ===")
                        ws = wb[sheet_name]
                        imported, skipped = import_sheet(ws, year, total_imported=total_imported)
                        print(f"  Imported: {imported}, Skipped: {skipped}")
                        total_imported += imported
                        total_skipped += skipped
                
                break  # Only use first file found
        
        print(f"\n=== IMPORT COMPLETE ===")
        print(f"Total Jobs Imported: {total_imported}")
        print(f"Total Skipped (duplicates): {total_skipped}")
        print(f"Total Customers: {Customer.query.count()}")
        print(f"Total Jobs: {Job.query.count()}")

if __name__ == '__main__':
    main()

