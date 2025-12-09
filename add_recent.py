#!/usr/bin/env python3
from openpyxl import load_workbook
from datetime import datetime, date
import sqlite3

def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if 'x' in value.lower():
            return None
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y']
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except:
                continue
    return None

def parse_price(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.replace('$', '').replace(',', '').strip()
        try:
            return float(value)
        except:
            return 0.0
    return 0.0

xlsx_path = "/home/bad/Desktop/David/quotes (version 1) (Autosaved) (Autosaved).xlsx"
db_path = "/home/bad/Desktop/David/quoteforge/instance/quoteforge.db"

wb = load_workbook(xlsx_path, data_only=True)
ws = wb['2014']

conn = sqlite3.connect(db_path)
cur = conn.cursor()

# Get next quote number
cur.execute("SELECT MAX(CAST(REPLACE(quote_number, 'Q', '') AS INTEGER)) FROM job")
next_num = (cur.fetchone()[0] or 0) + 1

imported = 0
for row_num, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
    row_date = row[1] if len(row) > 1 else None
    name = row[2] if len(row) > 2 else None
    address = row[3] if len(row) > 3 else None
    phone = row[4] if len(row) > 4 else None
    description = row[7] if len(row) > 7 else None
    price = row[9] if len(row) > 9 else 0
    
    if not name and not description:
        continue
        
    job_date = parse_date(row_date)
    if not job_date or job_date <= date(2022, 12, 14):
        continue
    
    name = str(name).strip() if name else "Unknown"
    phone = str(phone).strip() if phone and str(phone) != '0' else None
    address = str(address).strip() if address else None
    description = str(description).strip() if description else "No description"
    price_val = parse_price(price)
    
    # Find or create customer
    if phone:
        cur.execute("SELECT id FROM customer WHERE phone = ?", (phone,))
    else:
        cur.execute("SELECT id FROM customer WHERE name = ?", (name,))
    result = cur.fetchone()
    
    if result:
        customer_id = result[0]
    else:
        cur.execute("INSERT INTO customer (name, phone, address, created_at) VALUES (?, ?, ?, ?)",
                   (name, phone, address, datetime.now().isoformat()))
        customer_id = cur.lastrowid
    
    qn = f"Q{next_num:05d}"
    next_num += 1
    
    cur.execute("""INSERT INTO job (customer_id, quote_number, description, price, deposit, status, date, created_at)
                   VALUES (?, ?, ?, ?, 0, 'completed', ?, ?)""",
               (customer_id, qn, description, price_val, job_date.isoformat(), datetime.now().isoformat()))
    imported += 1

conn.commit()
conn.close()

print(f"Imported {imported} new jobs (2023-2025)")

