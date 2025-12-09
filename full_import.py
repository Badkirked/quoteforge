#!/usr/bin/env python3
"""
Full Import Script for QuoteForge
Imports ALL data from the master Excel file into a fresh database.
"""
from openpyxl import load_workbook
from datetime import datetime, date
import sqlite3
import os
import sys

# Configuration
XLSX_PATH = "/home/bad/Desktop/David/quotes (version 1) (Autosaved) (Autosaved).xlsx"
DB_PATH = "/home/bad/Desktop/David/quoteforge/instance/quoteforge.db"
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS customer (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name VARCHAR(200) NOT NULL,
    phone VARCHAR(50),
    email VARCHAR(200),
    address TEXT,
    created_at DATETIME
);

CREATE TABLE IF NOT EXISTS job (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER NOT NULL,
    quote_number VARCHAR(50) UNIQUE,
    description TEXT,
    price FLOAT,
    deposit FLOAT,
    status VARCHAR(50),
    notes TEXT,
    date DATE,
    created_at DATETIME,
    FOREIGN KEY(customer_id) REFERENCES customer(id)
);

CREATE TABLE IF NOT EXISTS backup (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    filename VARCHAR(200),
    created_at DATETIME,
    size INTEGER
);
"""

def parse_date(value):
    if value is None: return None
    if isinstance(value, (datetime, date)): 
        return value if isinstance(value, date) else value.date()
    if isinstance(value, str):
        value = value.strip()
        if not value or 'x' in value.lower(): return None
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y']:
            try: return datetime.strptime(value, fmt).date()
            except: continue
    return None

def parse_price(value):
    if value is None: return 0.0
    if isinstance(value, (int, float)): return float(value)
    if isinstance(value, str):
        val = value.replace('$', '').replace(',', '').strip()
        try: return float(val)
        except: return 0.0
    return 0.0

def main():
    print(f"Starting FULL IMPORT from: {XLSX_PATH}")
    
    # 1. Setup Database
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.executescript(SCHEMA_SQL)
    conn.commit()
    print("✓ Database created")

    # 2. Load Excel
    print("Loading Excel file (this may take a moment)...")
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb['2014']
    print(f"✓ Excel loaded. Total rows: {ws.max_row}")

    # 3. Import Data
    imported_jobs = 0
    customers_cache = {} # (name, phone) -> id
    
    # Start scanning from row 6
    print("Importing rows...")
    for row_num, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        # Data mapping: 1=Date, 2=Name, 3=Address, 4=Phone, 7=Desc, 8=Quote#, 9=Price
        
        # Validation
        if not row[2] and not row[7]: continue # Skip if no name AND no desc
        
        job_date = parse_date(row[1])
        if not job_date: job_date = date(2014, 1, 1) # Default fallback
        
        name = str(row[2]).strip() if row[2] else "Unknown"
        phone = str(row[4]).strip() if row[4] and str(row[4]) != '0' else None
        address = str(row[3]).strip() if row[3] else None
        desc = str(row[7]).strip() if row[7] else "No description"
        price = parse_price(row[9])
        
        # Quote Number
        qn_raw = row[8]
        if qn_raw:
            qn = f"Q{str(qn_raw).strip()}"
        else:
            qn = f"Q{imported_jobs + 1:05d}"

        # Customer Management
        cust_key = (name, phone)
        if cust_key in customers_cache:
            cust_id = customers_cache[cust_key]
        else:
            # Check DB
            if phone:
                cur.execute("SELECT id FROM customer WHERE phone = ?", (phone,))
            else:
                cur.execute("SELECT id FROM customer WHERE name = ?", (name,))
            
            res = cur.fetchone()
            if res:
                cust_id = res[0]
            else:
                cur.execute("INSERT INTO customer (name, phone, address, created_at) VALUES (?, ?, ?, ?)",
                           (name, phone, address, datetime.now().isoformat()))
                cust_id = cur.lastrowid
            
            customers_cache[cust_key] = cust_id

        # Insert Job
        try:
            cur.execute("""
                INSERT INTO job (customer_id, quote_number, description, price, deposit, status, date, created_at)
                VALUES (?, ?, ?, ?, 0, 'completed', ?, ?)
            """, (cust_id, qn, desc, price, job_date.isoformat(), datetime.now().isoformat()))
            imported_jobs += 1
        except sqlite3.IntegrityError:
            # Handle duplicate quote number by appending suffix
            qn = f"{qn}-{imported_jobs}"
            cur.execute("""
                INSERT INTO job (customer_id, quote_number, description, price, deposit, status, date, created_at)
                VALUES (?, ?, ?, ?, 0, 'completed', ?, ?)
            """, (cust_id, qn, desc, price, job_date.isoformat(), datetime.now().isoformat()))
            imported_jobs += 1

        if imported_jobs % 500 == 0:
            print(f"  Processed {imported_jobs} jobs...")
            conn.commit()

    conn.commit()
    
    # Verify
    cur.execute("SELECT COUNT(*) FROM job")
    total_jobs = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM customer")
    total_cust = cur.fetchone()[0]
    cur.execute("SELECT MAX(date) FROM job")
    max_date = cur.fetchone()[0]
    
    conn.close()
    
    print("\n=== IMPORT SUMMARY ===")
    print(f"Total Jobs: {total_jobs}")
    print(f"Total Customers: {total_cust}")
    print(f"Latest Job Date: {max_date}")
    print("======================")

if __name__ == '__main__':
    main()

